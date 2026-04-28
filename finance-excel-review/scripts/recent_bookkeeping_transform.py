#!/usr/bin/env python3
"""
Usage:
  python3 recent_bookkeeping_transform.py \
    --input-dir /home/che-an-wu/.openclaw/workspace/data/financial-templates/財務報表範例 \
    --output-dir /home/che-an-wu/.openclaw/workspace/data/financial-templates/output \
    [--year 2025] [--month 3]

Purpose:
- Read Chase Sapphire Reserve CSV, CTBC credit card CSV, and CTBC bank XLSX sample files.
- Produce three CSV outputs:
  1. normalized_rows.csv: unified source schema
  2. example4_rows.csv: initial example4-style bookkeeping rows
  3. example4_grouped_rows.csv: grouped rows closer to example4 same-name aggregation

This is an initial heuristic transformer, meant to be iterated as more real data and rules arrive.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
from collections import defaultdict
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Iterable, Optional


CHASE_MERCHANT_MAP = [
    (r"^COUPANG TAIWAN CO\. LTD$", "酷澎台灣官網Coupang"),
    (r"^CURSOR\s+AI POWERED IDE$", "Cursor"),
    (r"^GOOGLE \*Google One$", "Google One"),
    (r"^LINEPAY\*USPACETECH$", "USPACE"),
    (r"^LINEPAY\*AcerITSIncorpor$", "Acer ITS"),
    (r"^LINEPAY\*RTDigitalCOLTD$", "讀墨 Readmoo"),
    (r"^LINEPAY\*booksComCOLtd$", "博客來"),
    (r"^LINEPAY\*LouisacoffeeCOL$", "路易莎"),
    (r"^LINEPAY\*YOUPARKINGCOLTD$", "停車費"),
    (r"^LINEPAY\*TAIWANPARKINGCO$", "停車費"),
    (r"^LINE ?PAY\s*\*TAIPEI METRO$", "台北捷運"),
    (r"^TAIPEI METRO$", "台北捷運"),
    (r"^LINEPAY\*ZENSHO$", "ZENSHO"),
    (r"^LINE Pay Taiwan LimitedLI$", "LINE Pay"),
    (r"^UBER\s+\*BUSINESS$", "Uber"),
]

CTBC_BANK_COUNTERPARTY_MAP = {
    "GY": "Gy",
    "HU": "Hu",
    "LOV": "Lov",
    "街口": "街口",
}

CTBC_BANK_COMPANY_KEYWORDS = [
    "有限公司",
    "股份",
    "實見",
    "益循",
    "癒善糧",
    "法顧費用",
    "公司",
]

CTBC_BANK_NOISE_VALUES = {
    "-",
    "ＸＭＬ",
    "XML",
    "行動網",
    "網銀",
}

CTBC_TRANSFER_PURPOSE_PATTERNS = [
    (r"街口|JKOPAY", "街口儲值/轉帳", "待確認"),
    (r"信用卡款|卡費", "信用卡繳款", "其他"),
    (r"借款利息|利息", "借款利息", "其他"),
    (r"機車強制險|強制險|保險", "保險費", "其他"),
    (r"修烘乾機", "修烘乾機", "其他"),
    (r"隔熱紙和包膜修", "隔熱紙和包膜修", "其他"),
    (r"宜蘭住宿費|住宿", "住宿費", "行"),
    (r"馬偕紀念醫院|醫院|診所", "醫療費", "其他"),
    (r"法顧費用", "法顧費用", "公司往來"),
    (r"實見|益循|癒善糧|星夢農場|有限公司|股份", "公司往來", "公司往來"),
    (r"^Gy$", "健身課", "其他"),
    (r"^Hu$", "小虎上課學費", "其他"),
    (r"^Lov$", "Lov 轉帳", "待確認"),
]

from openpyxl import load_workbook


@dataclass
class NormalizedRow:
    source: str
    account: str
    tx_date: str
    post_date: str
    amount_twd: str
    direction: str
    merchant: str
    raw_description: str
    payment_channel: str
    order_id: str
    invoice_id: str
    category_hint: str
    note: str


@dataclass
class Example4Row:
    item: str
    amount: str
    category: str
    method: str
    day: str
    note: str
    source: str
    raw_item: str
    raw_category: str
    raw_amount: str
    tx_date: str


def parse_amount(value) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    text = text.replace("NTD", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def to_date_text(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    for fmt in ("%Y/%m/%d", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return text


def date_day(value: str) -> str:
    if not value:
        return ""
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", value)
    if m:
        return str(int(m.group(3)))
    return value


def year_month_matches(date_text: str, year: Optional[int], month: Optional[int]) -> bool:
    if not year and not month:
        return True
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", date_text or "")
    if not m:
        return False
    y = int(m.group(1))
    mo = int(m.group(2))
    if year and y != year:
        return False
    if month and mo != month:
        return False
    return True


def money_formula_twd_to_usd(amount: float) -> str:
    body = str(int(amount)) if float(amount).is_integer() else str(round(amount, 2))
    return f"=ROUND({body}/30,2)"


def twd_note(amount: float) -> str:
    y = round(amount / 30, 2)
    body = str(int(amount)) if float(amount).is_integer() else str(round(amount, 2))
    return f"{body}NTD {body}/30={y:.2f}"


def formula_sum(parts: list[str], method: str) -> str:
    if not parts:
        return ""
    if method == "chase":
        if len(parts) == 1:
            return parts[0]
        return "=" + "+ ".join(parts) + " "
    if len(parts) == 1:
        return money_formula_twd_to_usd(parse_amount(parts[0]) or 0)
    joined = "+".join(parts)
    return f"=ROUND(({joined})/30,2)"


def note_sum(parts: list[str]) -> str:
    if not parts:
        return ""
    if len(parts) == 1:
        amount = parse_amount(parts[0]) or 0
        return twd_note(amount)
    joined = "+".join(parts)
    total = sum(parse_amount(p) or 0 for p in parts)
    return f"{joined}NTD ({joined})/30={total/30:.2f}"


def chase_category_map(category: str) -> str:
    mapping = {
        "Food & Drink": "食",
        "Groceries": "食",
        "Travel": "行",
        "Automotive": "行",
        "Shopping": "奢",
        "Personal": "其他",
        "Entertainment": "奢",
    }
    return mapping.get((category or "").strip(), "其他")


def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def apply_pattern_map(text: str, patterns: list[tuple[str, str]]) -> str:
    normalized = normalize_spaces(text)
    for pattern, repl in patterns:
        if re.match(pattern, normalized, flags=re.IGNORECASE):
            return repl
    return normalized


def clean_chase_merchant(text: str) -> str:
    merchant = apply_pattern_map(text, CHASE_MERCHANT_MAP)
    merchant = re.sub(r"\s{2,}", " ", merchant).strip()
    return merchant


def clean_ctbc_card_merchant(text: str) -> str:
    t = (text or "").strip()
    t = re.sub(r"\s+TAIPEI\s+TW$", "", t)
    t = re.sub(r"\s+TW$", "", t)
    t = t.strip()
    replacements = [
        (r"^全聯.*", "全聯"),
        (r"^街口電支－統一超商.*", "街口支付－統一超商"),
        (r"^街口電支－全家便利商店.*", "街口支付－全家便利商店"),
        (r"^樂購蝦皮.*", "蝦皮"),
        (r"^eTag臨停.*", "停車費"),
        (r"^網銀行動繳.*", "網銀行動繳"),
        (r"^APPLE\.COM/BILL.*", "Apple"),
        (r"^電信.*", "電信費"),
        (r"^國外交易手續費.*", "國外交易手續費"),
    ]
    for pattern, repl in replacements:
        if re.match(pattern, t):
            return repl
    return t


def is_probable_bank_code(text: str) -> bool:
    t = normalize_spaces(text)
    return bool(re.fullmatch(r"\d{2,4}", t) or re.fullmatch(r"\d{3}/[0-9*\-]{6,}", t))


def is_probable_account_like(text: str) -> bool:
    t = normalize_spaces(text)
    return bool(re.fullmatch(r"[0-9*\-/]{6,}", t))


def is_noise_text(text: str) -> bool:
    t = normalize_spaces(text)
    return not t or t in CTBC_BANK_NOISE_VALUES or is_probable_bank_code(t) or is_probable_account_like(t)


def looks_like_real_transfer_label(text: str) -> bool:
    t = normalize_spaces(text)
    if is_noise_text(t):
        return False
    if re.search(r"(利息|信用卡款|保險|修|停車|房租|管理費|法顧費用|醫院|街口)", t):
        return True
    if any(k in t for k in CTBC_BANK_COMPANY_KEYWORDS):
        return True
    if re.search(r"[\u4e00-\u9fff]{2,}", t):
        return True
    if re.fullmatch(r"[A-Za-z][A-Za-z\s]{1,20}", t):
        return True
    return False


def normalize_ctbc_bank_counterparty(marker: str, note: str, counterparty_account: str) -> str:
    marker_clean = normalize_spaces(marker)
    note_clean = normalize_spaces(note)
    account_clean = normalize_spaces(counterparty_account)

    marker_mapped = CTBC_BANK_COUNTERPARTY_MAP.get(marker_clean.upper(), marker_clean) if marker_clean else ""
    if looks_like_real_transfer_label(marker_mapped):
        return marker_mapped

    for candidate in [note_clean, account_clean]:
        if looks_like_real_transfer_label(candidate):
            return candidate

    if marker_mapped and not is_noise_text(marker_mapped):
        return marker_mapped
    if note_clean and not is_noise_text(note_clean):
        return note_clean
    if account_clean and not is_noise_text(account_clean):
        return account_clean
    return ""


def apply_ctbc_transfer_purpose(item: str, summary: str, note: str, counterparty_account: str) -> tuple[str, str]:
    text = " ".join([normalize_spaces(item), normalize_spaces(summary), normalize_spaces(note), normalize_spaces(counterparty_account)])
    for pattern, mapped_item, mapped_category in CTBC_TRANSFER_PURPOSE_PATTERNS:
        if re.search(pattern, text, flags=re.IGNORECASE):
            return mapped_item, mapped_category
    return item, ""


def clean_ctbc_bank_item(summary: str, marker: str, note: str = "", counterparty_account: str = "") -> str:
    summary = normalize_spaces(summary)
    better_name = normalize_ctbc_bank_counterparty(marker, note, counterparty_account)
    if summary in {"轉帳提", "跨行轉", "現金提", "手續費"} and better_name:
        item = better_name
    elif summary == "現金提":
        item = "現金提款"
    else:
        mapping = {
            "跨行轉": "跨行轉帳",
            "轉帳提": "轉帳支出",
            "手續費": "手續費",
        }
        item = mapping.get(summary, better_name or summary)
    item = item if not is_noise_text(item) else summary
    mapped_item, _ = apply_ctbc_transfer_purpose(item, summary, note, counterparty_account)
    return mapped_item


def infer_ctbc_bank_category(item: str, summary: str, note: str, counterparty_account: str = "") -> str:
    mapped_item, mapped_category = apply_ctbc_transfer_purpose(item, summary, note, counterparty_account)
    if mapped_category:
        return mapped_category
    text = " ".join([mapped_item, summary, note, counterparty_account])
    if any(k in text for k in CTBC_BANK_COMPANY_KEYWORDS):
        return "公司往來"
    if any(k in text for k in ["停車", "捷運", "高鐵", "車", "機車", "油"]):
        return "行"
    if any(k in text for k in ["餐", "早餐", "午餐", "晚餐", "咖啡", "全聯", "壽司"]):
        return "食"
    if any(k in text for k in ["蝦皮", "Apple", "Coupang", "酷澎", "博客來", "讀墨", "展", "票"]):
        return "奢"
    if any(k in text for k in ["保險", "信用卡款", "貸款", "利息", "房租", "管理費", "醫院"]):
        return "其他"
    return "待確認" if summary in {"跨行轉", "轉帳提", "現金提"} else ""


def attach_ctbc_bank_fees(rows: list[Example4Row]) -> list[Example4Row]:
    result: list[Example4Row] = []
    pending_fee_rows: list[Example4Row] = []

    for row in rows:
        if row.source != "ctbc_bank":
            result.append(row)
            continue

        if row.raw_item == "手續費":
            pending_fee_rows.append(row)
            continue

        fees_to_attach = [
            fee for fee in pending_fee_rows
            if fee.tx_date == row.tx_date and abs((parse_amount(fee.raw_amount) or 0)) <= 30
        ]
        if fees_to_attach:
            base_amount = parse_amount(row.raw_amount) or 0
            total_fee = sum(parse_amount(fee.raw_amount) or 0 for fee in fees_to_attach)
            row.raw_amount = str(round(base_amount + total_fee, 2))
            fee_note = "+".join([str(int(parse_amount(f.raw_amount))) if (parse_amount(f.raw_amount) or 0).is_integer() else str(parse_amount(f.raw_amount)) for f in fees_to_attach])
            row.note = " | ".join([p for p in [row.note, f"含手續費 {fee_note} NTD"] if p])
            pending_fee_rows = [fee for fee in pending_fee_rows if fee not in fees_to_attach]
        result.append(row)

    result.extend(pending_fee_rows)
    return result


def ctbc_card_category(item: str) -> str:
    if any(k in item for k in ["全聯", "超商", "壽司", "餐", "家便利商店"]):
        return "食"
    if any(k in item for k in ["停車", "eTag", "悠遊卡", "交通"]):
        return "行"
    if any(k in item for k in ["蝦皮", "Apple"]):
        return "奢"
    return "其他"


def detect_direction(expense, income) -> tuple[str, Optional[float]]:
    exp = parse_amount(expense)
    inc = parse_amount(income)
    if exp not in (None, 0):
        return "expense", float(exp)
    if inc not in (None, 0):
        return "income", -float(inc)
    return "unknown", None


def read_chase(path: Path) -> tuple[list[NormalizedRow], list[Example4Row]]:
    normalized, example = [], []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            raw_amount = parse_amount(row.get("Amount"))
            if raw_amount is None:
                continue
            amount = raw_amount * -1
            tx_date = to_date_text(row.get("Transaction Date"))
            post_date = to_date_text(row.get("Post Date"))
            merchant = (row.get("Description") or "").strip()
            category = (row.get("Category") or "").strip()
            normalized.append(NormalizedRow(
                source="chase",
                account="Chase Sapphire Reserve",
                tx_date=tx_date,
                post_date=post_date,
                amount_twd=str(round(amount, 2)),
                direction="expense" if amount >= 0 else "income",
                merchant=clean_chase_merchant(merchant),
                raw_description=merchant,
                payment_channel="credit_card",
                order_id="",
                invoice_id="",
                category_hint=category,
                note=((row.get("Type") or "").strip() + (f"; {(row.get('Memo') or '').strip()}" if row.get("Memo") else "")).strip("; "),
            ))
            example.append(Example4Row(
                item=clean_chase_merchant(merchant),
                amount=str(round(amount, 2)),
                category=chase_category_map(category),
                method="chase",
                day=date_day(tx_date),
                note="",
                source="chase",
                raw_item=merchant,
                raw_category=category,
                raw_amount=str(round(amount, 2)),
                tx_date=tx_date,
            ))
    return normalized, example


def read_ctbc_card(path: Path) -> tuple[list[NormalizedRow], list[Example4Row]]:
    normalized, example = [], []
    text = path.read_bytes().decode("big5", errors="replace")
    lines = [line for line in text.splitlines() if line.strip()]
    header_idx = next(i for i, line in enumerate(lines) if "消費日" in line)
    reader = csv.DictReader(lines[header_idx:])
    for row in reader:
        amount = parse_amount(row.get("新臺幣金額"))
        if amount is None:
            continue
        tx_date = to_date_text(row.get("消費日"))
        post_date = to_date_text(row.get("入帳起息日") or row.get("入帳日"))
        raw_item = (row.get("摘要") or "").strip()
        item = clean_ctbc_card_merchant(raw_item)
        last4 = (row.get("末四碼") or row.get("卡號末四碼") or "").strip()
        normalized.append(NormalizedRow(
            source="ctbc_card",
            account=f"中信信用卡{last4}" if last4 else "中信信用卡",
            tx_date=tx_date,
            post_date=post_date,
            amount_twd=str(amount),
            direction="expense" if amount >= 0 else "transfer",
            merchant=item,
            raw_description=raw_item,
            payment_channel="credit_card",
            order_id=(row.get("交易序號") or "").strip(),
            invoice_id="",
            category_hint=ctbc_card_category(item),
            note="",
        ))
        category = "待確認" if amount < 0 or item in {"網銀行動繳"} else ctbc_card_category(item)
        example.append(Example4Row(
            item=item,
            amount=str(int(amount)) if float(amount).is_integer() else str(round(amount, 2)),
            category=category,
            method="中信信用卡",
            day=date_day(tx_date),
            note="",
            source="ctbc_card",
            raw_item=raw_item,
            raw_category=category,
            raw_amount=str(amount),
            tx_date=tx_date,
        ))
    return normalized, example


def read_ctbc_bank(path: Path) -> tuple[list[NormalizedRow], list[Example4Row]]:
    normalized, example = [], []
    wb = load_workbook(path, data_only=False)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(i for i, row in enumerate(rows) if row and row[0] == "日期")
    headers = [h if h is not None else "" for h in rows[header_idx]]
    for values in rows[header_idx + 1:]:
        if not any(v is not None and v != "" for v in values):
            continue
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        direction, amount = detect_direction(row.get("支出"), row.get("存入"))
        if amount is None:
            continue
        tx_date = to_date_text(row.get("日期"))
        summary = str(row.get("摘要") or "").strip()
        note = str(row.get("備註") or "").strip()
        marker = "" if row.get("註記") is None else str(row.get("註記")).strip()
        counterparty_account = str(row.get("對方帳號") or "").strip()
        item = clean_ctbc_bank_item(summary, marker, note, counterparty_account)
        merged_note = " | ".join([
            p for p in [note, counterparty_account, marker]
            if p and p != "None" and p != item and not is_noise_text(p)
        ])
        normalized.append(NormalizedRow(
            source="ctbc_bank",
            account="中信帳戶",
            tx_date=tx_date,
            post_date="",
            amount_twd=str(amount),
            direction=direction,
            merchant=item,
            raw_description=summary,
            payment_channel="bank_account",
            order_id="",
            invoice_id="",
            category_hint="",
            note=merged_note,
        ))
        category = infer_ctbc_bank_category(item, summary, merged_note, counterparty_account)
        example.append(Example4Row(
            item=item,
            amount=str(int(amount)) if float(amount).is_integer() else str(round(amount, 2)),
            category=category,
            method="中信活存",
            day=date_day(tx_date),
            note=merged_note,
            source="ctbc_bank",
            raw_item=summary,
            raw_category=category,
            raw_amount=str(amount),
            tx_date=tx_date,
        ))
    return normalized, example


def group_example4_rows(rows: list[Example4Row]) -> list[Example4Row]:
    grouped = defaultdict(list)
    for row in rows:
        key = (row.source, row.method, row.day, row.item, row.category)
        grouped[key].append(row)

    result: list[Example4Row] = []
    for key, members in grouped.items():
        source, method, day, item, category = key
        amount_parts = [m.raw_amount for m in members if m.raw_amount]
        if source == "chase":
            amount = formula_sum(amount_parts, method)
            note = members[0].note if len(members) == 1 else f"合併{len(members)}筆同名稱交易"
        else:
            amount = formula_sum(amount_parts, method)
            note = note_sum(amount_parts)
            if any(m.note for m in members):
                extras = [m.note for m in members if m.note]
                extra_text = " | ".join(dict.fromkeys(extras))
                note = f"{note} | {extra_text}" if note else extra_text
        raw_items = " | ".join(dict.fromkeys([m.raw_item for m in members if m.raw_item]))
        raw_categories = " | ".join(dict.fromkeys([m.raw_category for m in members if m.raw_category]))
        raw_amount = "+".join(amount_parts)
        result.append(Example4Row(
            item=item,
            amount=amount,
            category=category,
            method=method,
            day=day,
            note=note,
            source=source,
            raw_item=raw_items,
            raw_category=raw_categories,
            raw_amount=raw_amount,
            tx_date=members[0].tx_date,
        ))

    result.sort(key=lambda r: (int(r.day) if str(r.day).isdigit() else 99, r.method, r.item))
    return result


def write_csv(path: Path, rows: Iterable[dict]):
    rows = list(rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input-dir", required=True)
    ap.add_argument("--output-dir", required=True)
    ap.add_argument("--year", type=int)
    ap.add_argument("--month", type=int)
    args = ap.parse_args()

    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir)

    normalized_rows: list[NormalizedRow] = []
    example_rows: list[Example4Row] = []

    for path in sorted(input_dir.iterdir()):
        name = path.name
        if name == "Chase Sapphire Reserve 消費記錄.CSV":
            n, e = read_chase(path)
        elif name.startswith("中信卡") and path.suffix.lower() == ".csv":
            n, e = read_ctbc_card(path)
        elif name.startswith("中信帳戶") and path.suffix.lower() == ".xlsx":
            n, e = read_ctbc_bank(path)
        else:
            continue
        normalized_rows.extend(n)
        example_rows.extend(e)

    normalized_rows = [r for r in normalized_rows if year_month_matches(r.tx_date, args.year, args.month)]
    example_rows = [r for r in example_rows if year_month_matches(r.tx_date, args.year, args.month)]
    example_rows = attach_ctbc_bank_fees(example_rows)
    grouped_rows = group_example4_rows(example_rows)

    write_csv(output_dir / "normalized_rows.csv", [asdict(r) for r in normalized_rows])
    write_csv(output_dir / "example4_rows.csv", [asdict(r) for r in example_rows])
    write_csv(output_dir / "example4_grouped_rows.csv", [asdict(r) for r in grouped_rows])
    (output_dir / "summary.json").write_text(json.dumps({
        "normalized_count": len(normalized_rows),
        "example4_count": len(example_rows),
        "grouped_count": len(grouped_rows),
        "sources": sorted(list({r.source for r in normalized_rows})),
        "year": args.year,
        "month": args.month,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "normalized_count": len(normalized_rows),
        "example4_count": len(example_rows),
        "grouped_count": len(grouped_rows),
        "output_dir": str(output_dir),
        "year": args.year,
        "month": args.month,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
